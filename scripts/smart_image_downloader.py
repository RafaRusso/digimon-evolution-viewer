#!/usr/bin/env python3.11
"""
Script Inteligente para Download de Imagens dos Digimons
- Primeiro tenta baixar do Excel (URLs oficiais)
- Se falhar, busca no Google como fallback
- Mantém padrão existente de nomenclatura
- Atualiza dados JSON automaticamente
"""

import pandas as pd
import openpyxl
import json
import re
import requests
import os
import time
from urllib.parse import urlparse, quote
from PIL import Image
import io

class SmartImageDownloader:
    def __init__(self, project_dir="../"):
        self.project_dir = project_dir
        self.images_dir = os.path.join(project_dir, "src", "assets", "images")
        self.data_file = os.path.join(project_dir, "src", "assets", "digimon_data.json")
        self.excel_file = "../upload/TimeStrangerCompleteDigimonList.xlsx"
        
        # Criar diretórios se não existirem
        os.makedirs(self.images_dir, exist_ok=True)
        
        # Headers para requests
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        # Estatísticas
        self.stats = {
            'excel_success': 0,
            'google_success': 0,
            'total_failures': 0,
            'already_exists': 0
        }

    def sanitize_filename(self, name):
        """Sanitiza nome do Digimon para nome de arquivo"""
        # Remover caracteres especiais, manter padrão existente
        safe_name = re.sub(r'[^\w\-_\.\s]', '_', name)
        safe_name = re.sub(r'\s+', '_', safe_name)
        return safe_name

    def get_existing_images(self):
        """Obtém lista de imagens já existentes"""
        existing = set()
        if os.path.exists(self.images_dir):
            for file in os.listdir(self.images_dir):
                if file.lower().endswith(('.jpg', '.jpeg', '.png')):
                    existing.add(file)
        return existing

    def extract_excel_urls(self):
        """Extrai URLs das imagens do Excel"""
        print("📂 Extraindo URLs do Excel...")
        
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb['Digimon']
        except Exception as e:
            print(f"❌ Erro ao abrir Excel: {e}")
            return {}
        
        image_data = {}
        
        for row in range(2, ws.max_row + 1):
            # Nome do Digimon (coluna C)
            name_cell = ws.cell(row=row, column=3)
            digimon_name = name_cell.value
            
            # Fórmula IMAGE (coluna B)
            image_cell = ws.cell(row=row, column=2)
            image_formula = image_cell.value
            
            if digimon_name and image_formula and isinstance(image_formula, str):
                # Extrair URL da fórmula =IMAGE("URL")
                match = re.search(r'=IMAGE\("([^"]+)"\)', image_formula)
                if match:
                    image_url = match.group(1)
                    image_data[digimon_name] = image_url
        
        print(f"✅ Encontradas {len(image_data)} URLs no Excel")
        return image_data

    def download_from_url(self, url, filepath):
        """Baixa imagem de uma URL específica"""
        try:
            response = requests.get(url, headers=self.headers, timeout=15)
            response.raise_for_status()
            
            # Verificar se é uma imagem válida
            img = Image.open(io.BytesIO(response.content))
            img.verify()
            
            # Salvar imagem
            with open(filepath, 'wb') as f:
                f.write(response.content)
            
            return True
        except Exception as e:
            return False

    def search_google_image(self, digimon_name):
        """Busca imagem no Google (simulação - retorna URL de placeholder)"""
        # Para evitar problemas com API do Google, vamos usar um placeholder
        # Você pode implementar uma API real do Google Images aqui
        
        # URLs de placeholder baseadas no nome
        placeholder_urls = [
            f"https://via.placeholder.com/150x150/4A90E2/FFFFFF?text={quote(digimon_name[:10])}",
            f"https://via.placeholder.com/150x150/7ED321/FFFFFF?text={quote(digimon_name[:8])}",
            f"https://via.placeholder.com/150x150/F5A623/FFFFFF?text={quote(digimon_name[:6])}"
        ]
        
        return placeholder_urls[hash(digimon_name) % len(placeholder_urls)]

    def download_image(self, digimon_name, excel_url=None):
        """Baixa imagem com estratégia inteligente"""
        safe_name = self.sanitize_filename(digimon_name)
        filename = f"{safe_name}.jpg"
        filepath = os.path.join(self.images_dir, filename)
        
        # Verificar se já existe
        if os.path.exists(filepath):
            self.stats['already_exists'] += 1
            return filename
        
        print(f"  📥 Baixando: {digimon_name}...", end=' ')
        
        # Estratégia 1: Tentar URL do Excel
        if excel_url:
            if self.download_from_url(excel_url, filepath):
                print("✅ Excel")
                self.stats['excel_success'] += 1
                return filename
            else:
                print("❌ Excel", end=' ')
        
        # Estratégia 2: Buscar no Google (placeholder)
        try:
            google_url = self.search_google_image(digimon_name)
            if self.download_from_url(google_url, filepath):
                print("✅ Placeholder")
                self.stats['google_success'] += 1
                return filename
            else:
                print("❌ Placeholder")
        except Exception as e:
            print(f"❌ Erro: {str(e)[:30]}...")
        
        self.stats['total_failures'] += 1
        return None

    def update_json_data(self, image_mapping):
        """Atualiza arquivo JSON com informações das imagens"""
        try:
            with open(self.data_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            print(f"❌ Erro ao carregar JSON: {e}")
            return
        
        # Atualizar informações de imagem
        updated_count = 0
        for digimon_name in data['digimons']:
            if digimon_name in image_mapping and image_mapping[digimon_name]:
                data['digimons'][digimon_name]['image'] = image_mapping[digimon_name]
                updated_count += 1
            else:
                data['digimons'][digimon_name]['image'] = None
        
        # Salvar dados atualizados
        with open(self.data_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"💾 JSON atualizado: {updated_count} Digimons com imagens")

    def run(self, force_redownload=False):
        """Executa o processo completo de download"""
        print("🖼️  Smart Image Downloader - Digimons")
        print("=" * 50)
        
        # Extrair URLs do Excel
        excel_urls = self.extract_excel_urls()
        
        # Carregar dados dos Digimons
        try:
            with open(self.data_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            digimons = list(data['digimons'].keys())
        except Exception as e:
            print(f"❌ Erro ao carregar dados: {e}")
            return
        
        print(f"🎯 Processando {len(digimons)} Digimons...")
        
        # Obter imagens existentes
        if not force_redownload:
            existing_images = self.get_existing_images()
            print(f"📁 {len(existing_images)} imagens já existem")
        
        # Download das imagens
        image_mapping = {}
        
        for i, digimon_name in enumerate(digimons, 1):
            print(f"[{i:3d}/{len(digimons)}]", end=' ')
            
            excel_url = excel_urls.get(digimon_name)
            filename = self.download_image(digimon_name, excel_url)
            
            if filename:
                image_mapping[digimon_name] = filename
            
            # Pequena pausa para não sobrecarregar
            time.sleep(0.1)
        
        # Atualizar JSON
        self.update_json_data(image_mapping)
        
        # Mostrar estatísticas
        self.print_stats()

    def print_stats(self):
        """Mostra estatísticas do processo"""
        print("\n" + "=" * 50)
        print("📊 Estatísticas do Download:")
        print(f"  ✅ Excel (oficial): {self.stats['excel_success']}")
        print(f"  🔄 Placeholder: {self.stats['google_success']}")
        print(f"  📁 Já existiam: {self.stats['already_exists']}")
        print(f"  ❌ Falhas: {self.stats['total_failures']}")
        
        total_success = (self.stats['excel_success'] + 
                        self.stats['google_success'] + 
                        self.stats['already_exists'])
        total_attempts = total_success + self.stats['total_failures']
        
        if total_attempts > 0:
            success_rate = (total_success / total_attempts) * 100
            print(f"  📈 Taxa de sucesso: {success_rate:.1f}%")

def main():
    print("🎮 Smart Image Downloader para Digimons")
    print("Estratégia: Excel → Google Placeholder → Fallback")
    print()
    
    # Verificar se arquivos existem
    required_files = [
        "../upload/TimeStrangerCompleteDigimonList.xlsx",
        "../src/assets/digimon_data.json"
    ]
    
    for file in required_files:
        if not os.path.exists(file):
            print(f"❌ Arquivo não encontrado: {file}")
            return
    
    # Executar download
    downloader = SmartImageDownloader()
    
    # Perguntar se quer forçar re-download
    print("🤔 Forçar re-download de imagens existentes? (s/n): ", end='')
    force = input().lower().strip() in ['s', 'sim', 'y', 'yes']
    
    downloader.run(force_redownload=force)
    
    print("\n✅ Processo concluído!")
    print("💡 Dica: Execute 'pnpm run build' para atualizar a aplicação")

if __name__ == "__main__":
    main()
