#!/usr/bin/env python3.11
"""
Script Específico para Atualizar Apenas Imagens
- Foca só em baixar/atualizar imagens
- Não mexe nos dados dos Digimons
- Útil para executar periodicamente
- Mantém padrão de nomenclatura existente
"""

import pandas as pd
import openpyxl
import json
import re
import requests
import os
import time
from urllib.parse import urlparse
from PIL import Image
import io

class ImageUpdater:
    def __init__(self, project_dir="../"):
        self.project_dir = project_dir
        self.images_dir = os.path.join(project_dir, "src", "assets", "images")
        self.data_file = os.path.join(project_dir, "src", "assets", "digimon_data.json")
        self.excel_file = "../upload/TimeStrangerCompleteDigimonList.xlsx"
        
        # Headers para requests
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        
        # Estatísticas
        self.stats = {
            'downloaded': 0,
            'updated': 0,
            'failed': 0,
            'skipped': 0
        }

    def sanitize_filename(self, name):
        """Sanitiza nome para arquivo (mesmo padrão do projeto)"""
        safe_name = re.sub(r'[^\w\-_\.\s]', '_', name)
        safe_name = re.sub(r'\s+', '_', safe_name)
        return safe_name

    def get_excel_urls(self):
        """Extrai URLs do Excel"""
        print("📂 Lendo URLs do Excel...")
        
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb['Digimon']
        except Exception as e:
            print(f"❌ Erro ao abrir Excel: {e}")
            return {}
        
        urls = {}
        for row in range(2, ws.max_row + 1):
            name_cell = ws.cell(row=row, column=3)
            image_cell = ws.cell(row=row, column=2)
            
            digimon_name = name_cell.value
            image_formula = image_cell.value
            
            if digimon_name and image_formula and isinstance(image_formula, str):
                match = re.search(r'=IMAGE\("([^"]+)"\)', image_formula)
                if match:
                    urls[digimon_name] = match.group(1)
        
        print(f"✅ {len(urls)} URLs encontradas")
        return urls

    def download_image(self, url, filepath):
        """Baixa uma imagem específica"""
        try:
            response = requests.get(url, headers=self.headers, timeout=15)
            response.raise_for_status()
            
            # Verificar se é imagem válida
            img = Image.open(io.BytesIO(response.content))
            img.verify()
            
            # Salvar
            with open(filepath, 'wb') as f:
                f.write(response.content)
            
            return True
        except Exception:
            return False

    def update_missing_images(self, only_missing=True):
        """Atualiza imagens faltantes ou todas"""
        excel_urls = self.get_excel_urls()
        
        if not excel_urls:
            print("❌ Nenhuma URL encontrada no Excel")
            return
        
        print(f"🎯 Processando {len(excel_urls)} imagens...")
        
        for i, (digimon_name, url) in enumerate(excel_urls.items(), 1):
            safe_name = self.sanitize_filename(digimon_name)
            filename = f"{safe_name}.jpg"
            filepath = os.path.join(self.images_dir, filename)
            
            print(f"[{i:3d}/{len(excel_urls)}] {digimon_name}...", end=' ')
            
            # Verificar se já existe
            if only_missing and os.path.exists(filepath):
                print("⏭️  Já existe")
                self.stats['skipped'] += 1
                continue
            
            # Tentar baixar
            if self.download_image(url, filepath):
                if os.path.exists(filepath):
                    print("✅ Baixado")
                    self.stats['downloaded'] += 1
                else:
                    print("✅ Atualizado") 
                    self.stats['updated'] += 1
            else:
                print("❌ Falhou")
                self.stats['failed'] += 1
            
            # Pausa pequena
            time.sleep(0.1)

    def update_json_references(self):
        """Atualiza referências no JSON baseado nas imagens existentes"""
        print("\n💾 Atualizando referências no JSON...")
        
        try:
            with open(self.data_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            print(f"❌ Erro ao carregar JSON: {e}")
            return
        
        # Verificar imagens existentes
        existing_images = set()
        if os.path.exists(self.images_dir):
            existing_images = {f for f in os.listdir(self.images_dir) 
                             if f.lower().endswith(('.jpg', '.jpeg', '.png'))}
        
        # Atualizar dados
        updated = 0
        for digimon_name in data['digimons']:
            safe_name = self.sanitize_filename(digimon_name)
            possible_files = [
                f"{safe_name}.jpg",
                f"{safe_name}.jpeg", 
                f"{safe_name}.png"
            ]
            
            # Procurar arquivo existente
            found_file = None
            for possible_file in possible_files:
                if possible_file in existing_images:
                    found_file = possible_file
                    break
            
            # Atualizar referência
            old_image = data['digimons'][digimon_name].get('image')
            data['digimons'][digimon_name]['image'] = found_file
            
            if old_image != found_file:
                updated += 1
        
        # Salvar
        with open(self.data_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"✅ {updated} referências atualizadas no JSON")

    def clean_broken_images(self):
        """Remove imagens corrompidas"""
        print("\n🧹 Verificando imagens corrompidas...")
        
        if not os.path.exists(self.images_dir):
            return
        
        removed = 0
        for filename in os.listdir(self.images_dir):
            if not filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                continue
                
            filepath = os.path.join(self.images_dir, filename)
            try:
                with Image.open(filepath) as img:
                    img.verify()
            except Exception:
                print(f"  🗑️  Removendo: {filename}")
                os.remove(filepath)
                removed += 1
        
        if removed > 0:
            print(f"✅ {removed} imagens corrompidas removidas")
        else:
            print("✅ Nenhuma imagem corrompida encontrada")

    def print_stats(self):
        """Mostra estatísticas"""
        print("\n" + "=" * 40)
        print("📊 Estatísticas:")
        print(f"  📥 Baixadas: {self.stats['downloaded']}")
        print(f"  🔄 Atualizadas: {self.stats['updated']}")
        print(f"  ⏭️  Puladas: {self.stats['skipped']}")
        print(f"  ❌ Falhas: {self.stats['failed']}")

def main():
    print("🖼️  Atualizador de Imagens dos Digimons")
    print("=" * 40)
    
    # Verificar arquivos necessários
    updater = ImageUpdater()
    
    if not os.path.exists(updater.excel_file):
        print(f"❌ Excel não encontrado: {updater.excel_file}")
        return
    
    if not os.path.exists(updater.data_file):
        print(f"❌ JSON não encontrado: {updater.data_file}")
        return
    
    # Menu de opções
    print("\nOpções:")
    print("1. Baixar apenas imagens faltantes (recomendado)")
    print("2. Re-baixar todas as imagens")
    print("3. Apenas atualizar referências no JSON")
    print("4. Limpar imagens corrompidas")
    
    choice = input("\nEscolha uma opção (1-4): ").strip()
    
    if choice == "1":
        updater.update_missing_images(only_missing=True)
        updater.update_json_references()
    elif choice == "2":
        print("⚠️  Isso vai re-baixar TODAS as imagens!")
        confirm = input("Confirma? (s/n): ").lower().strip()
        if confirm in ['s', 'sim', 'y', 'yes']:
            updater.update_missing_images(only_missing=False)
            updater.update_json_references()
    elif choice == "3":
        updater.update_json_references()
    elif choice == "4":
        updater.clean_broken_images()
        updater.update_json_references()
    else:
        print("❌ Opção inválida")
        return
    
    updater.print_stats()
    print("\n✅ Processo concluído!")

if __name__ == "__main__":
    main()
