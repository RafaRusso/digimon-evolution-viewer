#!/usr/bin/env python3.11
import pandas as pd
import openpyxl
import json
import re
import requests
import os
from urllib.parse import urlparse

def find_excel_file():
    """
    Procura arquivo Excel na pasta data/
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, 'data')
    
    if not os.path.exists(data_dir):
        print(f"❌ Pasta 'data' não encontrada!")
        print(f"📁 Crie a pasta: {data_dir}")
        return None
    
    excel_files = []
    for file in os.listdir(data_dir):
        if file.endswith(('.xlsx', '.xls')):
            excel_files.append(os.path.join(data_dir, file))
    
    if not excel_files:
        print(f"❌ Nenhum arquivo Excel encontrado em: {data_dir}")
        return None
    
    return excel_files[0]  # Retorna o primeiro arquivo encontrado

def extract_image_urls(excel_path):
    """
    Extrai as URLs das imagens da planilha Excel.
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Digimon']
    
    image_data = {}
    
    print("🖼️  Extraindo URLs das imagens...")
    
    for row in range(2, ws.max_row + 1):  # Começar da linha 2 (pular cabeçalho)
        # Nome do Digimon (coluna C)
        name_cell = ws.cell(row=row, column=3)
        digimon_name = name_cell.value
        
        # Fórmula IMAGE (coluna B)
        image_cell = ws.cell(row=row, column=2)
        image_formula = image_cell.value
        
        if digimon_name and image_formula and isinstance(image_formula, str):
            # Extrair URL da fórmula =IMAGE("URL")
            match = re.search(r'=IMAGE\("([^"]+)"\)', image_formula)
            if match:
                image_url = match.group(1)
                image_data[digimon_name] = image_url
                print(f"  ✓ {digimon_name}")
    
    print(f"\n📊 Total de imagens encontradas: {len(image_data)}")
    return image_data

def download_images(image_data):
    """
    Baixa as imagens dos Digimons.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    download_dir = os.path.join(script_dir, 'src', 'assets', 'images')
    os.makedirs(download_dir, exist_ok=True)
    
    downloaded_images = {}
    failed_downloads = []
    
    print(f"\n⬇️  Baixando imagens para: src/assets/images/")
    
    for i, (digimon_name, image_url) in enumerate(image_data.items(), 1):
        try:
            print(f"  [{i:3d}/{len(image_data)}] {digimon_name}...", end=' ')
            
            # Fazer request da imagem
            response = requests.get(image_url, timeout=10)
            response.raise_for_status()
            
            # Determinar extensão do arquivo
            parsed_url = urlparse(image_url)
            file_extension = os.path.splitext(parsed_url.path)[1] or '.jpg'
            
            # Nome do arquivo (sanitizar nome do Digimon)
            safe_name = re.sub(r'[^\w\-_\.]', '_', digimon_name)
            filename = f"{safe_name}{file_extension}"
            filepath = os.path.join(download_dir, filename)
            
            # Salvar imagem
            with open(filepath, 'wb') as f:
                f.write(response.content)
            
            # Registrar sucesso
            downloaded_images[digimon_name] = filename
            print("✅")
            
        except Exception as e:
            print(f"❌ {str(e)[:50]}...")
            failed_downloads.append((digimon_name, str(e)))
    
    print(f"\n📈 Resultado do download:")
    print(f"  ✅ Sucessos: {len(downloaded_images)}")
    print(f"  ❌ Falhas: {len(failed_downloads)}")
    
    if failed_downloads and len(failed_downloads) <= 10:
        print(f"\n❌ Falhas:")
        for name, error in failed_downloads:
            print(f"  - {name}: {error[:60]}...")
    elif failed_downloads:
        print(f"\n❌ Primeiras 5 falhas:")
        for name, error in failed_downloads[:5]:
            print(f"  - {name}: {error[:60]}...")
    
    return downloaded_images, failed_downloads

def update_digimon_data(downloaded_images):
    """
    Atualiza o arquivo de dados dos Digimons com as informações das imagens.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_path = os.path.join(script_dir, 'src', 'assets', 'digimon_data.json')
    
    if not os.path.exists(data_path):
        print(f"❌ Arquivo de dados não encontrado: {data_path}")
        print("💡 Execute primeiro o script process_digimon_data.py")
        return
    
    # Carregar dados existentes
    with open(data_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Adicionar informações de imagem
    images_added = 0
    for digimon_name in data['digimons']:
        if digimon_name in downloaded_images:
            data['digimons'][digimon_name]['image'] = downloaded_images[digimon_name]
            images_added += 1
        else:
            data['digimons'][digimon_name]['image'] = None
    
    # Salvar dados atualizados
    with open(data_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"\n💾 Dados atualizados!")
    print(f"  📊 {images_added} Digimons agora têm imagens")
    print(f"  📁 Arquivo: src/assets/digimon_data.json")

if __name__ == "__main__":
    print("🖼️  Extrator de Imagens dos Digimons")
    print("=" * 50)
    
    # Encontrar arquivo Excel
    excel_path = find_excel_file()
    if not excel_path:
        exit()
    
    print(f"📂 Processando: {os.path.basename(excel_path)}")
    
    # Extrair URLs
    image_data = extract_image_urls(excel_path)
    
    if not image_data:
        print("❌ Nenhuma imagem encontrada para processar")
        exit()
    
    # Perguntar se quer continuar
    print(f"\n🤔 Deseja baixar {len(image_data)} imagens? (s/n): ", end='')
    response = input().lower().strip()
    
    if response not in ['s', 'sim', 'y', 'yes']:
        print("❌ Download cancelado pelo usuário")
        exit()
    
    # Baixar imagens
    downloaded_images, failed_downloads = download_images(image_data)
    
    if downloaded_images:
        # Atualizar dados
        update_digimon_data(downloaded_images)
        print("\n✅ Processo concluído!")
    else:
        print("\n❌ Nenhuma imagem foi baixada com sucesso")
    
    print(f"\n📁 Estrutura do projeto:")
    print(f"   digimon-evolution-viewer/")
    print(f"   ├── data/                    ← Excel aqui")
    print(f"   ├── src/assets/")
    print(f"   │   ├── images/              ← Imagens baixadas")
    print(f"   │   └── digimon_data.json   ← Dados com imagens")
    print(f"   └── extract_images.py       ← Este script")
