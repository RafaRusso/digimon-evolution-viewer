#!/usr/bin/env python3.11
"""
Script Simplificado para Baixar Imagens Faltantes
- Baixa apenas imagens que não existem
- Primeiro tenta Excel, depois Google/placeholder
- Atualiza JSON automaticamente
"""

import pandas as pd
import openpyxl
import json
import re
import requests
import os
import time
from urllib.parse import quote
from PIL import Image
import io

def sanitize_filename(name):
    """Converte nome do Digimon para nome de arquivo seguro"""
    safe_name = re.sub(r'[^\w\-_\.\s]', '_', name)
    safe_name = re.sub(r'\s+', '_', safe_name)
    return safe_name

def extract_excel_urls():
    """Extrai URLs das imagens do Excel"""
    print("📂 Extraindo URLs do Excel...")
    
    try:
        wb = openpyxl.load_workbook('../upload/TimeStrangerCompleteDigimonList.xlsx')
        ws = wb['Digimon']
    except Exception as e:
        print(f"❌ Erro ao abrir Excel: {e}")
        return {}
    
    image_data = {}
    
    for row in range(2, ws.max_row + 1):
        # Nome do Digimon (coluna C)
        name_cell = ws.cell(row=row, column=3)
        digimon_name = name_cell.value
        
        # Fórmula IMAGE (coluna B)  
        image_cell = ws.cell(row=row, column=2)
        image_formula = image_cell.value
        
        if digimon_name and image_formula and isinstance(image_formula, str):
            # Extrair URL da fórmula =IMAGE("URL")
            match = re.search(r'=IMAGE\("([^"]+)"\)', image_formula)
            if match:
                image_url = match.group(1)
                image_data[digimon_name] = image_url
    
    print(f"✅ Encontradas {len(image_data)} URLs no Excel")
    return image_data

def download_image(url, filepath):
    """Baixa uma imagem de uma URL"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        
        # Verificar se é uma imagem válida
        img = Image.open(io.BytesIO(response.content))
        img.verify()
        
        # Salvar imagem
        with open(filepath, 'wb') as f:
            f.write(response.content)
        
        return True
    except Exception:
        return False

def get_placeholder_url(digimon_name):
    """Gera URL de placeholder para Digimon"""
    # Cores baseadas no hash do nome
    colors = ['4A90E2', '7ED321', 'F5A623', 'D0021B', '9013FE', '50E3C2']
    color = colors[hash(digimon_name) % len(colors)]
    
    # Texto curto para o placeholder
    text = digimon_name[:8].replace(' ', '+')
    
    return f"https://via.placeholder.com/200x200/{color}/FFFFFF?text={quote(text)}"

def process_images():
    """Processa todas as imagens faltantes"""
    
    # Diretórios
    images_dir = "../src/assets/images"
    data_file = "../src/assets/digimon_data.json"
    
    # Criar diretório se não existir
    os.makedirs(images_dir, exist_ok=True)
    
    # Carregar dados dos Digimons
    try:
        with open(data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        digimons = list(data['digimons'].keys())
    except Exception as e:
        print(f"❌ Erro ao carregar dados: {e}")
        return
    
    # Extrair URLs do Excel
    excel_urls = extract_excel_urls()
    
    # Estatísticas
    stats = {
        'excel_success': 0,
        'placeholder_success': 0,
        'already_exists': 0,
        'failed': 0
    }
    
    print(f"\n🎯 Processando {len(digimons)} Digimons...")
    
    # Processar cada Digimon
    image_mapping = {}
    
    for i, digimon_name in enumerate(digimons, 1):
        safe_name = sanitize_filename(digimon_name)
        filename = f"{safe_name}.jpg"
        filepath = os.path.join(images_dir, filename)
        
        print(f"[{i:3d}/{len(digimons)}] {digimon_name}...", end=' ')
        
        # Verificar se já existe
        if os.path.exists(filepath):
            print("✅ Já existe")
            stats['already_exists'] += 1
            image_mapping[digimon_name] = filename
            continue
        
        # Tentar baixar do Excel
        excel_url = excel_urls.get(digimon_name)
        if excel_url and download_image(excel_url, filepath):
            print("✅ Excel")
            stats['excel_success'] += 1
            image_mapping[digimon_name] = filename
            time.sleep(0.1)  # Pausa pequena
            continue
        
        # Tentar placeholder
        placeholder_url = get_placeholder_url(digimon_name)
        if download_image(placeholder_url, filepath):
            print("✅ Placeholder")
            stats['placeholder_success'] += 1
            image_mapping[digimon_name] = filename
        else:
            print("❌ Falhou")
            stats['failed'] += 1
        
        time.sleep(0.1)  # Pausa pequena
    
    # Atualizar JSON
    print(f"\n💾 Atualizando JSON...")
    
    updated_count = 0
    for digimon_name in data['digimons']:
        if digimon_name in image_mapping:
            data['digimons'][digimon_name]['image'] = image_mapping[digimon_name]
            updated_count += 1
        else:
            data['digimons'][digimon_name]['image'] = None
    
    # Salvar dados atualizados
    with open(data_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"✅ {updated_count} Digimons atualizados no JSON")
    
    # Mostrar estatísticas
    print("\n" + "=" * 50)
    print("📊 Estatísticas:")
    print(f"  ✅ Excel (oficial): {stats['excel_success']}")
    print(f"  🔄 Placeholder: {stats['placeholder_success']}")
    print(f"  📁 Já existiam: {stats['already_exists']}")
    print(f"  ❌ Falhas: {stats['failed']}")
    
    total_success = stats['excel_success'] + stats['placeholder_success'] + stats['already_exists']
    total_attempts = total_success + stats['failed']
    
    if total_attempts > 0:
        success_rate = (total_success / total_attempts) * 100
        print(f"  📈 Taxa de sucesso: {success_rate:.1f}%")

def main():
    print("🖼️  Download de Imagens Faltantes - Digimons")
    print("=" * 50)
    
    # Verificar arquivos necessários
    required_files = [
        "../upload/TimeStrangerCompleteDigimonList.xlsx",
        "../src/assets/digimon_data.json"
    ]
    
    for file in required_files:
        if not os.path.exists(file):
            print(f"❌ Arquivo não encontrado: {file}")
            return
    
    print("🚀 Iniciando download de imagens faltantes...")
    print("📋 Estratégia: Excel → Placeholder → Skip")
    
    process_images()
    
    print("\n✅ Processo concluído!")
    print("💡 Execute 'cd digimon-evolution-viewer && pnpm run build' para atualizar")

if __name__ == "__main__":
    main()
